import argparse
import concurrent.futures
import html
import json
import shutil
import re
import subprocess
import sys
import tempfile
from email.message import EmailMessage
from collections import Counter
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Set, Tuple
from zipfile import ZIP_DEFLATED, ZipFile

from lxml import etree
from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string


PLACEHOLDER_RE = re.compile(r"\{\{([A-Z0-9_]+)\}\}")
WORD_NAMESPACE = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
NSMAP = {"w": WORD_NAMESPACE}
XML_DECLARATION = '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
DEFAULT_XML_PART_PREFIXES = (
    "word/document.xml",
    "word/header",
    "word/footer",
    "word/footnotes.xml",
    "word/endnotes.xml",
)
PROGRESS_PREFIX = "__GREEKLIT_PROGRESS__"
OUTLOOK_MSG_TIMEOUT_SECONDS = 30


@dataclass
class Config:
    workbook_path: Path
    worksheet_name: str
    contract_template_path: Path
    email_template_path: Path
    mapping_path: Path
    output_dir: Path
    data_start_row: int
    header_row: int
    date_format: str
    filename_pattern: str
    combined_email_filename: str
    email_output_mode: Optional[str]
    report_filename: str
    contract_output_subdir: str
    pdf_output_subdir: str
    email_output_subdir: str
    convert_to_pdf: bool
    pdf_conversion_workers: int
    keep_docx_output: bool
    attach_contract_to_eml: bool
    libreoffice_path: Optional[Path]
    row_identity_placeholders: List[str]
    skip_if_column_contains: Dict[str, List[str]]
    skip_if_column_equals: Dict[str, List[str]]
    skip_if_row_fill_colors: List[str]


@dataclass
class RowResult:
    row_number: int
    status: str
    contract_filename: Optional[str] = None
    pdf_filename: Optional[str] = None
    missing_placeholders: Optional[List[str]] = None
    skip_reason: Optional[str] = None


@dataclass
class PendingGeneratedRow:
    row_number: int
    contract_output_path: Path
    email_text: str
    contract_filename: Optional[str]
    row_values: Dict[str, str]


def emit_progress(
    stage: str,
    message: str,
    current: int = 0,
    total: int = 0,
    percent: Optional[int] = None,
    **metrics: int,
) -> None:
    payload = {
        "current": current,
        "message": message,
        "stage": stage,
        "total": total,
    }
    if percent is not None:
        payload["percent"] = max(0, min(100, int(percent)))
    for key, value in metrics.items():
        payload[key] = value
    print(f"{PROGRESS_PREFIX}{json.dumps(payload, ensure_ascii=True)}", flush=True)


def stage_percent(start: int, end: int, current: int, total: int) -> int:
    if total <= 0:
        return start
    return start + round(((end - start) * current) / total)


def load_config(path: Path) -> Config:
    raw = json.loads(path.read_text(encoding="utf-8"))
    base = path.parent

    def resolve(value: str) -> Path:
        candidate = Path(value)
        return candidate if candidate.is_absolute() else (base / candidate).resolve()

    return Config(
        workbook_path=resolve(raw["workbook_path"]),
        worksheet_name=raw["worksheet_name"],
        contract_template_path=resolve(raw["contract_template_path"]),
        email_template_path=resolve(raw["email_template_path"]),
        mapping_path=resolve(raw["mapping_path"]),
        output_dir=resolve(raw["output_dir"]),
        data_start_row=int(raw.get("data_start_row", 2)),
        header_row=int(raw.get("header_row", 1)),
        date_format=raw.get("date_format", "%Y-%m-%d"),
        filename_pattern=raw["filename_pattern"],
        combined_email_filename=raw.get("combined_email_filename", "email_drafts.html"),
        email_output_mode=raw.get("email_output_mode", "combined_docx"),
        report_filename=raw.get("report_filename", "generation_report.txt"),
        contract_output_subdir=raw.get("contract_output_subdir", "contracts"),
        pdf_output_subdir=raw.get("pdf_output_subdir", "contracts_pdf"),
        email_output_subdir=raw.get("email_output_subdir", "emails"),
        convert_to_pdf=bool(raw.get("convert_to_pdf", True)),
        pdf_conversion_workers=max(1, int(raw.get("pdf_conversion_workers", 2))),
        keep_docx_output=bool(raw.get("keep_docx_output", True)),
        attach_contract_to_eml=bool(raw.get("attach_contract_to_eml", True)),
        libreoffice_path=resolve(raw["libreoffice_path"]) if raw.get("libreoffice_path") else None,
        row_identity_placeholders=[str(item) for item in raw.get("row_identity_placeholders", [])],
        skip_if_column_contains={
            str(column).upper(): [str(item) for item in values]
            for column, values in raw.get("skip_if_column_contains", {}).items()
        },
        skip_if_column_equals={
            str(column).upper(): [str(item) for item in values]
            for column, values in raw.get("skip_if_column_equals", {}).items()
        },
        skip_if_row_fill_colors=[str(item).upper() for item in raw.get("skip_if_row_fill_colors", [])],
    )


def load_mapping(path: Path) -> Dict[str, str]:
    mapping: Dict[str, str] = {}
    for line_number, raw_line in enumerate(path.read_text(encoding="utf-8").splitlines(), start=1):
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue
        if "=" not in line:
            raise ValueError(f"Invalid mapping line {line_number}: {raw_line!r}")
        key, column = [part.strip() for part in line.split("=", 1)]
        if not key or not re.fullmatch(r"[A-Z0-9_]+", key):
            raise ValueError(f"Invalid placeholder name on line {line_number}: {key!r}")
        if not column or not re.fullmatch(r"[A-Z]+", column):
            raise ValueError(f"Invalid Excel column on line {line_number}: {column!r}")
        mapping[key] = column
    if not mapping:
        raise ValueError("Mapping file is empty.")
    return mapping


def extract_docx_parts_and_placeholders(path: Path) -> Tuple[Dict[str, bytes], Set[str], Set[str]]:
    placeholders: Set[str] = set()
    xml_parts: Set[str] = set()
    with ZipFile(path, "r") as zf:
        for name in zf.namelist():
            if not any(name == prefix or name.startswith(prefix) for prefix in DEFAULT_XML_PART_PREFIXES):
                continue
            xml_parts.add(name)
            try:
                data = zf.read(name)
            except KeyError:
                continue
            placeholders.update(PLACEHOLDER_RE.findall(data.decode("utf-8", errors="ignore")))
    return {}, placeholders, xml_parts


def extract_text_placeholders(text: str) -> Set[str]:
    return set(PLACEHOLDER_RE.findall(text))


def validate_paths(config: Config) -> None:
    for path in (
        config.workbook_path,
        config.contract_template_path,
        config.email_template_path,
        config.mapping_path,
    ):
        if not path.exists():
            raise FileNotFoundError(f"Required file not found: {path}")


def get_header_value(ws, column_letter: str, header_row: int) -> str:
    value = ws[f"{column_letter}{header_row}"].value
    return normalize_cell_value(value, "%Y-%m-%d")


def normalize_cell_value(value, date_format: str) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, datetime):
        return value.strftime(date_format)
    if isinstance(value, date):
        return value.strftime(date_format)
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        normalized = format(Decimal(str(value)).normalize(), "f")
        return normalized.rstrip("0").rstrip(".") if "." in normalized else normalized
    return str(value).strip()


def sanitize_filename(value: str) -> str:
    cleaned = re.sub(r'[<>:"/\\|?*]', "-", value)
    cleaned = re.sub(r"\s+", " ", cleaned).strip(" .")
    return cleaned or "contract"


def normalize_color_code(color_value) -> str:
    if color_value is None:
        return ""
    return str(color_value).upper()


def render_plain_text(template: str, values: Dict[str, str]) -> str:
    def replacer(match: re.Match[str]) -> str:
        return values.get(match.group(1), match.group(0))

    return PLACEHOLDER_RE.sub(replacer, template)

def render_html_template(template: str, values: Dict[str, str]) -> str:
    def replacer(match: re.Match[str]) -> str:
        replacement = values.get(match.group(1))
        return html.escape(replacement) if replacement is not None else match.group(0)

    return PLACEHOLDER_RE.sub(replacer, template)


def write_combined_email_drafts_html(path: Path, drafts: Sequence[Tuple[str, str]]) -> None:
    sections: List[str] = []

    for heading, rendered_html in drafts:
        sections.append(
            "\n".join(
                [
                    '<section style="margin-bottom: 28px;">',
                    f"<p><strong>{html.escape(heading)}</strong></p>",
                    rendered_html,
                    "</section>",
                ]
            )
        )

    document_html = "\n".join(
        [
            "<!DOCTYPE html>",
            '<html lang="en">',
            "<head>",
            '<meta charset="utf-8" />',
            "<title>Email Drafts</title>",
            "</head>",
            '<body style="font-family: Calibri, Arial, sans-serif; font-size: 11pt; line-height: 1.45;">',
            "\n".join(sections) if sections else "<p>No email drafts were generated.</p>",
            "</body>",
            "</html>",
        ]
    )

    path.write_text(document_html, encoding="utf-8")


def wrap_single_email_html(rendered_html: str, heading: str) -> str:
    return "\n".join(
        [
            "<!DOCTYPE html>",
            '<html lang="en">',
            "<head>",
            '<meta charset="utf-8" />',
            f"<title>{html.escape(heading)}</title>",
            "</head>",
            '<body style="font-family: Calibri, Arial, sans-serif; font-size: 11pt; line-height: 1.45;">',
            rendered_html,
            "</body>",
            "</html>",
        ]
    )


def _extract_email_headers(rendered_html: str) -> Dict[str, str]:
    """Extract Subject/To/Cc from rendered HTML header paragraphs."""
    def clean(value: str) -> Optional[str]:
        if not value:
            return None
        unescaped = html.unescape(value.strip())
        if "undefined" in unescaped.lower():
            cleaned = re.sub(r"\s*<\s*undefined\s*>\s*", "", unescaped, flags=re.IGNORECASE)
            cleaned = re.sub(r"\bundefined\b", "", cleaned, flags=re.IGNORECASE).strip()
            return cleaned or None
        return unescaped or None

    headers: Dict[str, str] = {}
    for field in ("Subject", "To", "Cc"):
        m = re.search(rf"<p><strong>{field}:</strong>\s*([^<]*?)\s*</p>", rendered_html)
        if m:
            val = clean(m.group(1))
            if val:
                headers[field.lower()] = val
    return headers


def _strip_email_headers(rendered_html: str) -> str:
    """Remove Subject/To/Cc header <p> elements from rendered HTML body."""
    return re.sub(
        r"<p><strong>(?:Subject|To|Cc):</strong>\s*[^<]*</p>",
        "",
        rendered_html,
    )


def _sanitize_recipient_token(token: str) -> Optional[str]:
    value = html.unescape((token or "").strip())
    if not value:
        return None
    value = re.sub(r"\s*<?\s*undefined\s*>?\s*", "", value, flags=re.IGNORECASE).strip()
    if not value:
        return None
    if "{{" in value and "}}" in value:
        return None
    return value


def _split_recipient_addresses(raw_value: str) -> List[str]:
    parts = re.split(r"[;,]", raw_value or "")
    cleaned: List[str] = []
    for part in parts:
        candidate = _sanitize_recipient_token(part)
        if candidate:
            cleaned.append(candidate)
    return cleaned


def write_eml_draft(path: Path, row_values: Dict[str, str], rendered_html: str, row_number: int) -> None:
    message = EmailMessage()
    headers = _extract_email_headers(rendered_html)
    body_html = _strip_email_headers(rendered_html)
    subject = headers.get("subject") or row_values.get("TITLE") or row_values.get("APPLICATION_CODE") or f"Generated draft row {row_number}"
    message["Subject"] = subject
    to_value = headers.get("to") or row_values.get("EMAIL_TO", "")
    cc_value = headers.get("cc") or row_values.get("EMAIL_CC", "")
    to_addresses = _split_recipient_addresses(to_value)
    cc_addresses = _split_recipient_addresses(cc_value)
    if to_addresses:
        message["To"] = "; ".join(to_addresses)
    if cc_addresses:
        message["Cc"] = "; ".join(cc_addresses)
    message.set_content("This is an HTML email draft generated by Doc Gen Studio.")
    message.add_alternative(body_html, subtype="html")
    path.write_text(message.as_string(), encoding="utf-8")


def create_msg_draft_via_outlook(
    path: Path,
    row_values: Dict[str, str],
    rendered_html: str,
    row_number: int,
    attachment_path: Optional[Path] = None,
    report_stage=None,
) -> None:
    def stage(value: str) -> None:
        if report_stage:
            report_stage(value)

    try:
        stage("import pythoncom/win32com")
        import pythoncom  # type: ignore
        import win32com.client  # type: ignore
    except Exception as exc:
        raise RuntimeError("Outlook MSG output requires pywin32 and Microsoft Outlook on Windows.") from exc

    stage("CoInitialize")
    pythoncom.CoInitialize()
    outlook = None
    mail = None
    try:
        stage("Dispatch Outlook.Application")
        outlook = win32com.client.Dispatch("Outlook.Application")
        stage("CreateItem")
        mail = outlook.CreateItem(0)
        stage("set fields")
        headers = _extract_email_headers(rendered_html)
        body_html = _strip_email_headers(rendered_html)
        mail.Subject = headers.get("subject") or row_values.get("TITLE") or row_values.get("APPLICATION_CODE") or f"Generated draft row {row_number}"
        to_value = headers.get("to") or row_values.get("EMAIL_TO", "")
        cc_value = headers.get("cc") or row_values.get("EMAIL_CC", "")
        to_addresses = _split_recipient_addresses(to_value)
        cc_addresses = _split_recipient_addresses(cc_value)

        if to_addresses:
            mail.To = "; ".join(to_addresses)
        if cc_addresses:
            mail.CC = "; ".join(cc_addresses)

        try:
            mail.Recipients.ResolveAll()
        except Exception:
            pass
        mail.HTMLBody = body_html
        if attachment_path and attachment_path.exists():
            stage("add attachment")
            mail.Attachments.Add(str(attachment_path.resolve()))
        stage("SaveAs")
        mail.SaveAs(str(path.resolve()), 9)
        stage("saved")
    except Exception as exc:
        raise RuntimeError(f"Could not create Outlook MSG draft: {exc}") from exc
    finally:
        mail = None
        outlook = None
        stage("CoUninitialize")
        pythoncom.CoUninitialize()


def self_worker_command() -> List[str]:
    if getattr(sys, "frozen", False):
        return [sys.executable]
    return [sys.executable, str(Path(__file__).resolve())]


def run_outlook_msg_worker(payload_path: Path) -> int:
    payload = json.loads(payload_path.read_text(encoding="utf-8"))
    stage_path = Path(payload["stage_path"])

    def report_stage(value: str) -> None:
        stage_path.write_text(value, encoding="utf-8")

    create_msg_draft_via_outlook(
        Path(payload["path"]),
        {str(key): str(value) for key, value in payload["row_values"].items()},
        str(payload["rendered_html"]),
        int(payload["row_number"]),
        Path(payload["attachment_path"]) if payload.get("attachment_path") else None,
        report_stage,
    )
    return 0


def write_msg_draft(
    path: Path,
    row_values: Dict[str, str],
    rendered_html: str,
    row_number: int,
    attachment_path: Optional[Path] = None,
) -> None:
    with tempfile.TemporaryDirectory(prefix="docgen-msg-") as temp_dir:
        temp_path = Path(temp_dir)
        payload_path = temp_path / "payload.json"
        stage_path = temp_path / "stage.txt"
        payload = {
            "attachment_path": str(attachment_path.resolve()) if attachment_path else None,
            "path": str(path.resolve()),
            "rendered_html": rendered_html,
            "row_number": row_number,
            "row_values": row_values,
            "stage_path": str(stage_path),
        }
        payload_path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")

        command = [*self_worker_command(), "--outlook-msg-worker", str(payload_path)]
        try:
            result = subprocess.run(
                command,
                capture_output=True,
                text=True,
                timeout=OUTLOOK_MSG_TIMEOUT_SECONDS,
                check=False,
            )
        except subprocess.TimeoutExpired as exc:
            last_stage = stage_path.read_text(encoding="utf-8") if stage_path.exists() else "startup"
            raise RuntimeError(
                f"Timed out after {OUTLOOK_MSG_TIMEOUT_SECONDS}s while creating Outlook MSG draft "
                f"for row {row_number} ({last_stage}). Close Outlook completely, reopen it, and try again."
            ) from exc

        if result.returncode != 0:
            details = (result.stderr or result.stdout).strip()
            raise RuntimeError(details or f"Outlook MSG worker exited with code {result.returncode}.")


def collect_used_placeholders(*placeholder_sets: Iterable[str]) -> Set[str]:
    result: Set[str] = set()
    for items in placeholder_sets:
        result.update(items)
    return result


def relevant_xml_part(name: str) -> bool:
    return any(name == prefix or name.startswith(prefix) for prefix in DEFAULT_XML_PART_PREFIXES)


def replace_placeholders_in_docx(template_path: Path, output_path: Path, values: Dict[str, str]) -> None:
    with ZipFile(template_path, "r") as source, ZipFile(output_path, "w", ZIP_DEFLATED) as target:
        for item in source.infolist():
            data = source.read(item.filename)
            if relevant_xml_part(item.filename) and item.filename.endswith(".xml"):
                data = replace_placeholders_in_xml_part(data, values)
            target.writestr(item, data)


def replace_placeholders_in_xml_part(xml_bytes: bytes, values: Dict[str, str]) -> bytes:
    parser = etree.XMLParser(remove_blank_text=False, recover=True)
    root = etree.fromstring(xml_bytes, parser=parser)

    for paragraph in root.xpath(".//w:p", namespaces=NSMAP):
        replace_placeholders_in_container(paragraph, values)

    for table_cell in root.xpath(".//w:tc", namespaces=NSMAP):
        if not table_cell.xpath("./w:p", namespaces=NSMAP):
            replace_placeholders_in_container(table_cell, values)

    rendered = etree.tostring(root, encoding="UTF-8", xml_declaration=False)
    return (XML_DECLARATION.encode("utf-8") + rendered) if not rendered.startswith(b"<?xml") else rendered


def replace_placeholders_in_container(container, values: Dict[str, str]) -> None:
    text_nodes = container.xpath(".//w:t", namespaces=NSMAP)
    if not text_nodes:
        return

    original_text = "".join(node.text or "" for node in text_nodes)
    replaced_text = render_plain_text(original_text, values)
    if replaced_text == original_text:
        return

    first = text_nodes[0]
    first.text = replaced_text
    for node in text_nodes[1:]:
        node.text = ""


def ensure_unique_filename(base_name: str, extension: str, seen: Counter, row_number: int) -> str:
    stem = sanitize_filename(base_name)
    seen[stem] += 1
    suffix = f" - row {row_number}" if seen[stem] > 1 else ""
    return f"{stem}{suffix}{extension}"


def build_row_values(ws, row_number: int, mapping: Dict[str, str], date_format: str) -> Dict[str, str]:
    values: Dict[str, str] = {}
    for placeholder, column_letter in mapping.items():
        cell_value = ws[f"{column_letter}{row_number}"].value
        values[placeholder] = normalize_cell_value(cell_value, date_format)
    return values


def row_has_any_content(values: Dict[str, str]) -> bool:
    return any(value != "" for value in values.values())


def row_has_identity(values: Dict[str, str], identity_placeholders: Sequence[str]) -> bool:
    if not identity_placeholders:
        return True
    return any(values.get(name, "") != "" for name in identity_placeholders)


def row_matches_skip_rules(ws, row_number: int, config: Config) -> Optional[str]:
    for column_letter, rejected_values in config.skip_if_column_equals.items():
        value = normalize_cell_value(ws[f"{column_letter}{row_number}"].value, config.date_format)
        if any(value == rejected_value for rejected_value in rejected_values):
            return f"column {column_letter} equals rejection value ({value})"

    for column_letter, keywords in config.skip_if_column_contains.items():
        value = normalize_cell_value(ws[f"{column_letter}{row_number}"].value, config.date_format)
        lowered = value.lower()
        if lowered and any(keyword.lower() in lowered for keyword in keywords):
            return f"column {column_letter} contains rejection marker ({value})"

    if config.skip_if_row_fill_colors:
        for cell in ws[row_number]:
            fill = cell.fill
            color = ""
            if fill and fill.fgColor:
                color = normalize_color_code(fill.fgColor.rgb or fill.fgColor.indexed or fill.fgColor.type)
            if color in config.skip_if_row_fill_colors:
                return f"row fill color marked as rejected ({color})"

    return None


def resolve_libreoffice_path(config: Config) -> Optional[Path]:
    candidates: List[Path] = []

    def add_candidate(path: Optional[Path]) -> None:
        if not path:
            return
        if path.suffix.lower() == ".exe":
            com_candidate = path.with_suffix(".com")
            candidates.append(com_candidate)
        candidates.append(path)

    if config.libreoffice_path:
        add_candidate(config.libreoffice_path)

    path_candidate = shutil.which("soffice.com") or shutil.which("soffice") or shutil.which("libreoffice")
    if path_candidate:
        add_candidate(Path(path_candidate))

    candidates.extend(
        [
            Path(r"C:\Program Files\LibreOffice\program\soffice.com"),
            Path(r"C:\Program Files\LibreOffice\program\soffice.exe"),
            Path(r"C:\Program Files (x86)\LibreOffice\program\soffice.com"),
            Path(r"C:\Program Files (x86)\LibreOffice\program\soffice.exe"),
        ]
    )

    seen: Set[Path] = set()
    for candidate in candidates:
        if candidate in seen:
            continue
        seen.add(candidate)
        if candidate and candidate.exists():
            return candidate
    return None


def check_word_pdf_capability() -> Tuple[bool, Optional[str]]:
    try:
        import win32com.client  # type: ignore
    except Exception:
        return False, "win32com is not available."

    word = None
    try:
        word = win32com.client.DispatchEx("Word.Application")
        return True, None
    except Exception as exc:
        return False, f"Microsoft Word COM automation is unavailable: {exc}"
    finally:
        if word is not None:
            try:
                word.Quit()
            except Exception:
                pass


def convert_docx_to_pdf_libreoffice(soffice_path: Path, docx_path: Path, pdf_dir: Path) -> Path:
    pdf_dir.mkdir(parents=True, exist_ok=True)
    with tempfile.TemporaryDirectory(prefix="greeklit-libreoffice-") as profile_dir:
        command = [
            str(soffice_path),
            f"-env:UserInstallation={Path(profile_dir).resolve().as_uri()}",
            "--convert-to",
            "pdf:writer_pdf_Export",
            "--outdir",
            str(pdf_dir),
            str(docx_path),
        ]
        completed = subprocess.run(command, capture_output=True, text=True, check=False)
    if completed.returncode != 0:
        raise RuntimeError(
            "LibreOffice PDF conversion failed: "
            + (completed.stderr.strip() or completed.stdout.strip() or f"exit code {completed.returncode}")
        )
    pdf_path = pdf_dir / f"{docx_path.stem}.pdf"
    if not pdf_path.exists():
        raise RuntimeError(f"LibreOffice reported success but no PDF was created for {docx_path.name}")
    return pdf_path


def convert_docx_to_pdf_word(docx_path: Path, pdf_path: Path) -> None:
    import win32com.client  # type: ignore

    word = win32com.client.DispatchEx("Word.Application")
    try:
        word.Visible = False
        document = word.Documents.Open(str(docx_path.resolve()))
        try:
            document.SaveAs(str(pdf_path.resolve()), FileFormat=17)
        finally:
            document.Close(False)
    finally:
        word.Quit()


def convert_docx_to_pdf(
    pdf_backend: str,
    soffice_path: Optional[Path],
    docx_path: Path,
    pdf_dir: Path,
) -> Path:
    if pdf_backend == "libreoffice" and soffice_path is not None:
        return convert_docx_to_pdf_libreoffice(soffice_path, docx_path, pdf_dir)
    if pdf_backend == "word":
        pdf_path = pdf_dir / docx_path.with_suffix(".pdf").name
        convert_docx_to_pdf_word(docx_path, pdf_path)
        return pdf_path
    raise RuntimeError("PDF conversion backend was not resolved.")


def write_report(
    path: Path,
    config: Config,
    mapping: Dict[str, str],
    header_snapshot: Dict[str, str],
    contract_placeholders: Sequence[str],
    email_placeholders: Sequence[str],
    unmapped_placeholders: Sequence[str],
    unused_mappings: Sequence[str],
    results: Sequence[RowResult],
) -> None:
    generated = [result for result in results if result.status == "generated"]
    skipped = [result for result in results if result.status == "skipped"]
    lines = [
        "Contract and Email Generation Report",
        "",
        f"Workbook: {config.workbook_path}",
        f"Worksheet: {config.worksheet_name}",
        f"Contract template: {config.contract_template_path}",
        f"Email template: {config.email_template_path}",
        f"Mapping file: {config.mapping_path}",
        f"Output directory: {config.output_dir}",
        "",
        "Header snapshot by mapped column:",
    ]
    for placeholder in sorted(mapping):
        lines.append(f"- {placeholder} = {mapping[placeholder]} (header: {header_snapshot.get(placeholder, '')})")

    rejection_rules = [
        f"{column} equals {', '.join(values)}"
        for column, values in sorted(config.skip_if_column_equals.items())
    ]

    lines.extend(
        [
            "",
            f"Contract placeholders: {', '.join(sorted(contract_placeholders)) or '(none)'}",
            f"Email placeholders: {', '.join(sorted(email_placeholders)) or '(none)'}",
            f"Unmapped placeholders: {', '.join(sorted(unmapped_placeholders)) or '(none)'}",
            f"Unused mappings: {', '.join(sorted(unused_mappings)) or '(none)'}",
            f"Rejection rules: {'; '.join(rejection_rules) if rejection_rules else '(none)'}",
            "",
            f"Generated rows: {len(generated)}",
            f"Skipped rows: {len(skipped)}",
            "",
            "Generated outputs:",
        ]
    )
    if generated:
        for result in generated:
            details = []
            if result.contract_filename:
                details.append(f"docx={result.contract_filename}")
            if result.pdf_filename:
                details.append(f"pdf={result.pdf_filename}")
            lines.append(f"- Row {result.row_number}: {', '.join(details)}")
    else:
        lines.append("- None")

    lines.append("")
    lines.append("Skipped rows:")
    if skipped:
        for result in skipped:
            if result.missing_placeholders:
                missing = ", ".join(result.missing_placeholders or [])
                lines.append(f"- Row {result.row_number}: missing {missing}")
            elif result.skip_reason:
                lines.append(f"- Row {result.row_number}: {result.skip_reason}")
            else:
                lines.append(f"- Row {result.row_number}: skipped")
    else:
        lines.append("- None")

    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate filled Word contracts and combined email drafts from an Excel workbook.",
    )
    parser.add_argument(
        "--config",
        default="generator_config.json",
        help="Path to the generator configuration JSON file.",
    )
    parser.add_argument(
        "--outlook-msg-worker",
        help=argparse.SUPPRESS,
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    if args.outlook_msg_worker:
        return run_outlook_msg_worker(Path(args.outlook_msg_worker))

    config_path = Path(args.config).resolve()
    emit_progress("prepare", "Loading generation configuration.", percent=2)
    config = load_config(config_path)
    validate_paths(config)
    mapping = load_mapping(config.mapping_path)

    emit_progress("prepare", "Inspecting templates and workbook.", percent=6)
    email_template = config.email_template_path.read_text(encoding="utf-8")
    email_output_mode = config.email_output_mode or "combined_docx"
    writes_visible_email_drafts = email_output_mode in {"combined_docx", "separate_docx", "separate_eml", "separate_msg", "separate_msg_with_docx", "separate_msg_with_pdf"}
    _, contract_placeholders, _ = extract_docx_parts_and_placeholders(config.contract_template_path)
    email_placeholders = extract_text_placeholders(email_template) if writes_visible_email_drafts else set()
    filename_placeholders = extract_text_placeholders(config.filename_pattern)
    used_placeholders = collect_used_placeholders(
        contract_placeholders,
        email_placeholders,
        filename_placeholders,
        config.row_identity_placeholders,
    )
    unmapped_placeholders = sorted(used_placeholders - set(mapping))
    unused_mappings = sorted(set(mapping) - used_placeholders)

    workbook = load_workbook(config.workbook_path, data_only=True)
    worksheet_name = (config.worksheet_name or "").strip()
    if not worksheet_name:
        worksheet_name = workbook.sheetnames[0]
    if worksheet_name not in workbook.sheetnames:
        raise ValueError(
            f"Worksheet {worksheet_name!r} not found. Available sheets: {', '.join(workbook.sheetnames)}"
        )
    ws = workbook[worksheet_name]
    total_rows = max(0, ws.max_row - config.data_start_row + 1)
    emit_progress(
        "prepare",
        f"{total_rows} workbook rows found.",
        total=total_rows,
        percent=8,
        rowsFound=total_rows,
    )

    max_column = ws.max_column
    for placeholder, column_letter in mapping.items():
        index = column_index_from_string(column_letter)
        if index > max_column:
            raise ValueError(
                f"Mapped column {column_letter} for placeholder {placeholder} is outside the worksheet range "
                f"(max column is {max_column})."
            )

    header_snapshot = {
        placeholder: get_header_value(ws, column_letter, config.header_row)
        for placeholder, column_letter in mapping.items()
    }

    config.output_dir.mkdir(parents=True, exist_ok=True)
    contracts_dir = config.output_dir / config.contract_output_subdir
    pdf_dir = config.output_dir / config.pdf_output_subdir
    contracts_dir.mkdir(parents=True, exist_ok=True)
    pdf_dir.mkdir(parents=True, exist_ok=True)
    emit_progress("prepare", "Output folders are ready.", percent=10)

    results: List[RowResult] = []
    pending_generated_rows: List[PendingGeneratedRow] = []
    filename_counter: Counter = Counter()
    docx_count = 0
    generated_count = 0
    skipped_count = 0
    pdf_available = False
    pdf_warning: Optional[str] = None
    pdf_backend: Optional[str] = None
    soffice_path = resolve_libreoffice_path(config)
    if config.convert_to_pdf:
        emit_progress("pdf-backend", "Checking PDF conversion backend.", percent=12)
        if soffice_path:
            pdf_available = True
            pdf_backend = "libreoffice"
        else:
            pdf_available, pdf_warning = check_word_pdf_capability()
            if pdf_available:
                pdf_backend = "word"

    pdf_runtime_failed = False
    for row_index, row_number in enumerate(range(config.data_start_row, ws.max_row + 1), start=1):
        emit_progress(
            "docx",
            f"Processing workbook row {row_number}.",
            current=row_index,
            total=total_rows,
            percent=stage_percent(15, 60, row_index, total_rows),
            docxCount=docx_count,
            generatedCount=generated_count,
            rowsFound=total_rows,
            skippedCount=skipped_count,
        )
        skip_reason = row_matches_skip_rules(ws, row_number, config)
        if skip_reason:
            results.append(RowResult(row_number=row_number, status="skipped", skip_reason=skip_reason))
            skipped_count += 1
            emit_progress(
                "docx",
                f"Skipped workbook row {row_number}.",
                current=row_index,
                total=total_rows,
                percent=stage_percent(15, 60, row_index, total_rows),
                docxCount=docx_count,
                generatedCount=generated_count,
                rowsFound=total_rows,
                skippedCount=skipped_count,
            )
            continue

        row_values = build_row_values(ws, row_number, mapping, config.date_format)
        if not row_has_any_content(row_values):
            continue
        if not row_has_identity(row_values, config.row_identity_placeholders):
            continue

        missing = sorted(name for name in used_placeholders if row_values.get(name, "") == "")
        if missing:
            results.append(RowResult(row_number=row_number, status="skipped", missing_placeholders=missing))
            skipped_count += 1
            emit_progress(
                "docx",
                f"Skipped workbook row {row_number}.",
                current=row_index,
                total=total_rows,
                percent=stage_percent(15, 60, row_index, total_rows),
                docxCount=docx_count,
                generatedCount=generated_count,
                rowsFound=total_rows,
                skippedCount=skipped_count,
            )
            continue

        contract_base_name = render_plain_text(config.filename_pattern, row_values)
        base_stem = sanitize_filename(contract_base_name)
        contract_filename = ensure_unique_filename(base_stem, ".docx", filename_counter, row_number)
        contract_output_path = contracts_dir / contract_filename
        replace_placeholders_in_docx(config.contract_template_path, contract_output_path, row_values)

        email_text = render_html_template(email_template, row_values)
        pending_generated_rows.append(
            PendingGeneratedRow(
                row_number=row_number,
                contract_output_path=contract_output_path,
                email_text=email_text,
                contract_filename=contract_filename,
                row_values=row_values,
            )
        )
        generated_count += 1
        docx_count = generated_count if config.keep_docx_output else 0
        emit_progress(
            "docx",
            f"Created DOCX for row {row_number}.",
            current=row_index,
            total=total_rows,
            percent=stage_percent(15, 60, row_index, total_rows),
            docxCount=docx_count,
            generatedCount=generated_count,
            rowsFound=total_rows,
            skippedCount=skipped_count,
        )

    pdf_paths_by_row: Dict[int, Path] = {}
    expected_docx_count = len(pending_generated_rows) if config.keep_docx_output else 0
    expected_email_count = len(pending_generated_rows) if writes_visible_email_drafts else 0
    expected_pdf_count = len(pending_generated_rows) if config.convert_to_pdf and pdf_available and pdf_backend else 0
    if config.convert_to_pdf and pdf_available and pdf_backend:
        total_pdfs = len(pending_generated_rows)
        emit_progress(
            "pdf",
            "Starting PDF conversion.",
            current=0,
            total=total_pdfs,
            percent=62,
            docxCount=expected_docx_count,
            emailDraftCount=0,
            expectedDocxCount=expected_docx_count,
            expectedEmailDraftCount=expected_email_count,
            expectedPdfCount=expected_pdf_count,
            generatedCount=generated_count,
            pdfCount=0,
            rowsFound=total_rows,
            skippedCount=skipped_count,
        )
        try:
            if pdf_backend == "libreoffice":
                max_workers = min(config.pdf_conversion_workers, max(1, len(pending_generated_rows)))
                with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
                    future_to_row = {
                        executor.submit(
                            convert_docx_to_pdf,
                            pdf_backend,
                            soffice_path,
                            pending.contract_output_path,
                            pdf_dir,
                        ): pending.row_number
                        for pending in pending_generated_rows
                    }
                    for completed_count, future in enumerate(concurrent.futures.as_completed(future_to_row), start=1):
                        row_number = future_to_row[future]
                        pdf_paths_by_row[row_number] = future.result()
                        emit_progress(
                            "pdf",
                            f"Converted PDF for row {row_number}.",
                            current=completed_count,
                            total=total_pdfs,
                            percent=stage_percent(62, 84, completed_count, total_pdfs),
                            docxCount=expected_docx_count,
                            emailDraftCount=0,
                            expectedDocxCount=expected_docx_count,
                            expectedEmailDraftCount=expected_email_count,
                            expectedPdfCount=expected_pdf_count,
                            generatedCount=generated_count,
                            pdfCount=completed_count,
                            rowsFound=total_rows,
                            skippedCount=skipped_count,
                        )
            else:
                for completed_count, pending in enumerate(pending_generated_rows, start=1):
                    pdf_paths_by_row[pending.row_number] = convert_docx_to_pdf(
                        pdf_backend,
                        soffice_path,
                        pending.contract_output_path,
                        pdf_dir,
                    )
                    emit_progress(
                        "pdf",
                        f"Converted PDF for row {pending.row_number}.",
                        current=completed_count,
                        total=total_pdfs,
                        percent=stage_percent(62, 84, completed_count, total_pdfs),
                        docxCount=expected_docx_count,
                        emailDraftCount=0,
                        expectedDocxCount=expected_docx_count,
                        expectedEmailDraftCount=expected_email_count,
                        expectedPdfCount=expected_pdf_count,
                        generatedCount=generated_count,
                        pdfCount=completed_count,
                        rowsFound=total_rows,
                        skippedCount=skipped_count,
                    )
        except Exception as exc:
            pdf_runtime_failed = True
            pdf_warning = f"PDF conversion failed at runtime: {exc}"
            pdf_paths_by_row = {}

    actual_pdf_count = len(pdf_paths_by_row)
    emit_progress(
        "finalize",
        "Finalizing generated output records.",
        percent=86,
        docxCount=expected_docx_count,
        emailDraftCount=0,
        expectedDocxCount=expected_docx_count,
        expectedEmailDraftCount=expected_email_count,
        expectedPdfCount=expected_pdf_count,
        generatedCount=generated_count,
        pdfCount=actual_pdf_count,
        rowsFound=total_rows,
        skippedCount=skipped_count,
    )
    for pending in pending_generated_rows:
        pdf_output_path = pdf_paths_by_row.get(pending.row_number)
        pdf_filename = pdf_output_path.name if pdf_output_path else None

        contract_filename = pending.contract_filename
        if not config.keep_docx_output and pending.contract_output_path.exists():
            pending.contract_output_path.unlink()
            contract_filename = None

        results.append(
            RowResult(
                row_number=pending.row_number,
                status="generated",
                contract_filename=contract_filename,
                pdf_filename=pdf_filename,
            )
        )

    results.sort(key=lambda item: item.row_number)

    combined_email_path = config.output_dir / config.combined_email_filename
    combined_email_docx_entries = []
    email_output_dir = config.output_dir / config.email_output_subdir
    if writes_visible_email_drafts:
        email_output_dir.mkdir(parents=True, exist_ok=True)
    email_filename_counter: Counter = Counter()
    total_email_entries = len(pending_generated_rows)
    for email_index, pending in enumerate(pending_generated_rows, start=1):
        emit_progress(
            "email" if writes_visible_email_drafts else "finalize",
            (
                f"Preparing email draft for row {pending.row_number}."
                if writes_visible_email_drafts
                else f"Preparing output summary for row {pending.row_number}."
            ),
            current=email_index,
            total=total_email_entries,
            percent=stage_percent(88, 94, email_index, total_email_entries),
            docxCount=expected_docx_count,
            emailDraftCount=email_index if writes_visible_email_drafts else 0,
            expectedDocxCount=expected_docx_count,
            expectedEmailDraftCount=expected_email_count,
            expectedPdfCount=expected_pdf_count,
            generatedCount=generated_count,
            pdfCount=actual_pdf_count,
            rowsFound=total_rows,
            skippedCount=skipped_count,
        )
        application_code = build_row_values(
            ws,
            pending.row_number,
            {"APPLICATION_CODE": mapping["APPLICATION_CODE"]} if "APPLICATION_CODE" in mapping else {},
            config.date_format,
        ).get("APPLICATION_CODE", "")
        heading = (
            f"===== {application_code} - ROW {pending.row_number} ====="
            if application_code
            else f"===== ROW {pending.row_number} ====="
        )
        if email_output_mode == "combined_docx":
            combined_email_docx_entries.append((heading, pending.email_text))
        elif email_output_mode == "separate_docx":
            draft_base_name = (
                f"email draft - {application_code}"
                if application_code
                else f"email draft row {pending.row_number}"
            )
            html_filename = ensure_unique_filename(draft_base_name, ".html", email_filename_counter, pending.row_number)
            html_path = email_output_dir / html_filename
            html_path.write_text(wrap_single_email_html(pending.email_text, heading), encoding="utf-8")
        elif email_output_mode == "separate_eml":
            draft_base_name = (
                f"email draft - {application_code}"
                if application_code
                else f"email draft row {pending.row_number}"
            )
            eml_filename = ensure_unique_filename(draft_base_name, ".eml", email_filename_counter, pending.row_number)
            eml_path = email_output_dir / eml_filename
            write_eml_draft(eml_path, pending.row_values, pending.email_text, pending.row_number)
        elif email_output_mode in ("separate_msg", "separate_msg_with_docx", "separate_msg_with_pdf"):
            draft_base_name = (
                f"email draft - {application_code}"
                if application_code
                else f"email draft row {pending.row_number}"
            )
            msg_filename = ensure_unique_filename(draft_base_name, ".msg", email_filename_counter, pending.row_number)
            msg_path = email_output_dir / msg_filename
            if email_output_mode == "separate_msg_with_docx":
                attachment_path = pending.contract_output_path if pending.contract_output_path.exists() else None
            elif email_output_mode == "separate_msg_with_pdf":
                attachment_path = pdf_paths_by_row.get(pending.row_number)
            else:
                attachment_path = pending.contract_output_path if config.attach_contract_to_eml else None
            write_msg_draft(msg_path, pending.row_values, pending.email_text, pending.row_number, attachment_path)

    if email_output_mode == "combined_docx":
        write_combined_email_drafts_html(combined_email_path, combined_email_docx_entries)

    emit_progress(
        "report",
        "Writing generation report.",
        percent=96,
        docxCount=expected_docx_count,
        emailDraftCount=expected_email_count,
        expectedDocxCount=expected_docx_count,
        expectedEmailDraftCount=expected_email_count,
        expectedPdfCount=expected_pdf_count,
        generatedCount=generated_count,
        pdfCount=actual_pdf_count,
        rowsFound=total_rows,
        skippedCount=skipped_count,
    )
    report_path = config.output_dir / config.report_filename
    write_report(
        report_path,
        config,
        mapping,
        header_snapshot,
        sorted(contract_placeholders),
        sorted(email_placeholders),
        unmapped_placeholders,
        unused_mappings,
        results,
    )

    print(f"Generated {len([r for r in results if r.status == 'generated'])} contracts.")
    print(f"Skipped {len([r for r in results if r.status == 'skipped'])} rows.")
    if email_output_mode == "combined_docx":
        print(f"Combined email drafts file: {combined_email_path}")
    elif writes_visible_email_drafts:
        print(f"Email drafts directory: {email_output_dir}")
    print(f"Contract DOCX directory: {contracts_dir}")
    print(f"Contract PDF directory: {pdf_dir}")
    print(f"Report: {report_path}")
    if config.convert_to_pdf and pdf_available and pdf_backend:
        print(f"PDF backend: {pdf_backend}")
        if pdf_backend == "libreoffice":
            print(f"PDF conversion workers: {min(config.pdf_conversion_workers, max(1, len(pending_generated_rows)))}")
    if config.convert_to_pdf and not pdf_available and pdf_warning:
        print(f"Warning: PDF conversion disabled at runtime. {pdf_warning}")
    if config.convert_to_pdf and pdf_runtime_failed and pdf_warning:
        print(f"Warning: {pdf_warning}")
    if unmapped_placeholders:
        print("Warning: unmapped placeholders found:", ", ".join(unmapped_placeholders))
    if unused_mappings:
        print("Warning: unused mappings found:", ", ".join(unused_mappings))
    emit_progress(
        "complete",
        "Generation complete.",
        percent=100,
        docxCount=expected_docx_count,
        emailDraftCount=expected_email_count,
        expectedDocxCount=expected_docx_count,
        expectedEmailDraftCount=expected_email_count,
        expectedPdfCount=expected_pdf_count,
        generatedCount=generated_count,
        pdfCount=actual_pdf_count,
        rowsFound=total_rows,
        skippedCount=skipped_count,
    )
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(f"Error: {exc}", file=sys.stderr)
        raise
