import json
import html
import re
import subprocess
import sys
import tempfile
from collections import Counter
from datetime import date, datetime
from decimal import Decimal
from email.message import EmailMessage
from pathlib import Path

from openpyxl import load_workbook

PLACEHOLDER_RE = re.compile(r"\{\{([A-Z0-9_]+)\}\}")
PROGRESS_PREFIX = "__GREEKLIT_PROGRESS__"
OUTLOOK_MSG_TIMEOUT_SECONDS = 30


def emit_progress(stage, message, current=0, total=0, percent=None, **metrics):
    payload = {
        "current": current,
        "message": message,
        "stage": stage,
        "total": total,
    }
    if percent is not None:
        payload["percent"] = max(0, min(100, int(percent)))
    for key, value in metrics.items():
        payload[key] = value
    print(f"{PROGRESS_PREFIX}{json.dumps(payload, ensure_ascii=True)}", flush=True)


def stage_percent(start, end, current, total):
    if total <= 0:
        return start
    return start + round(((end - start) * current) / total)


def normalize_cell_value(value):
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        normalized = format(Decimal(str(value)).normalize(), "f")
        return normalized.rstrip("0").rstrip(".") if "." in normalized else normalized
    return str(value).strip()


def render_template(template_text, values):
    def replacer(match):
        token = match.group(1)
        replacement = values.get(token)
        return html.escape(replacement) if replacement is not None else match.group(0)

    return PLACEHOLDER_RE.sub(replacer, template_text)


def extract_email_headers_from_rendered_html(rendered_html):
    """Extract Subject, To, and Cc from rendered HTML email template.
    
    Looks for patterns like:
    <p><strong>Subject:</strong> Some Subject Text</p>
    <p><strong>To:</strong> email@example.com</p>
    <p><strong>Cc:</strong> cc@example.com</p>
    
    Filters out 'undefined' strings that may result from unresolved placeholders.
    """
    def clean_value(value):
        """Remove HTML entities and filter out undefined/invalid values."""
        if not value:
            return None
        unescaped = html.unescape(value.strip())
        # Remove any 'undefined' string or values containing it
        if 'undefined' in unescaped.lower():
            # Remove the word undefined and its surrounding angle brackets
            cleaned = re.sub(r'\s*<?\s*undefined\s*>?\s*', '', unescaped, flags=re.IGNORECASE)
            if not cleaned.strip():
                return None
            return cleaned.strip()
        return unescaped if unescaped else None
    
    headers = {}
    
    # Extract Subject
    subject_match = re.search(r'<p><strong>Subject:</strong>\s*([^<]*?)\s*</p>', rendered_html)
    if subject_match:
        subject_value = clean_value(subject_match.group(1))
        if subject_value:
            headers['subject'] = subject_value
    
    # Extract To
    to_match = re.search(r'<p><strong>To:</strong>\s*([^<]*?)\s*</p>', rendered_html)
    if to_match:
        to_value = clean_value(to_match.group(1))
        if to_value:
            headers['to'] = to_value
    
    # Extract Cc
    cc_match = re.search(r'<p><strong>Cc:</strong>\s*([^<]*?)\s*</p>', rendered_html)
    if cc_match:
        cc_value = clean_value(cc_match.group(1))
        if cc_value:
            headers['cc'] = cc_value
    
    return headers


def strip_email_headers_from_html(rendered_html):
    """Remove Subject, To, and Cc header paragraphs from HTML.
    
    Returns the HTML with these header lines removed, leaving only the body content.
    """
    # Remove the Subject, To, and Cc header paragraphs
    cleaned = re.sub(
        r'<p><strong>(?:Subject|To|Cc):</strong>\s*[^<]*</p>',
        '',
        rendered_html
    )
    return cleaned


def build_row_values(worksheet, row_number, mapping):
    values = {}
    for token, column in mapping.items():
        values[token] = normalize_cell_value(worksheet[f"{column}{row_number}"].value)
    return values


def row_matches_skip_rules(worksheet, row_number, skip_if_column_equals):
    for column_letter, rejected_values in skip_if_column_equals.items():
        value = normalize_cell_value(worksheet[f"{column_letter}{row_number}"].value)
        if any(value == rejected_value for rejected_value in rejected_values):
            return True
    return False


def write_email_drafts_html(path, drafts):
    sections = []

    for heading, rendered_html in drafts:
        sections.append(
            "\n".join(
                [
                    '<section style="margin-bottom: 28px;">',
                    f"<p><strong>{html.escape(heading)}</strong></p>",
                    rendered_html,
                    "</section>",
                ]
            )
        )

    document_html = "\n".join(
        [
            "<!DOCTYPE html>",
            '<html lang="en">',
            "<head>",
            '<meta charset="utf-8" />',
            "<title>Email Drafts</title>",
            "</head>",
            '<body style="font-family: Calibri, Arial, sans-serif; font-size: 11pt; line-height: 1.45;">',
            "\n".join(sections) if sections else "<p>No email drafts were generated.</p>",
            "</body>",
            "</html>",
        ]
    )

    path.write_text(document_html, encoding="utf-8")


def sanitize_filename(value):
    cleaned = re.sub(r'[<>:"/\\|?*]', "-", value)
    cleaned = re.sub(r"\s+", " ", cleaned).strip(" .")
    return cleaned or "email draft"


def ensure_unique_filename(base_name, extension, seen, row_number):
    stem = sanitize_filename(base_name)
    seen[stem] += 1
    suffix = f" - row {row_number}" if seen[stem] > 1 else ""
    return f"{stem}{suffix}{extension}"


def wrap_single_email_html(rendered_html, heading):
    return "\n".join(
        [
            "<!DOCTYPE html>",
            '<html lang="en">',
            "<head>",
            '<meta charset="utf-8" />',
            f"<title>{html.escape(heading)}</title>",
            "</head>",
            '<body style="font-family: Calibri, Arial, sans-serif; font-size: 11pt; line-height: 1.45;">',
            rendered_html,
            "</body>",
            "</html>",
        ]
    )


def write_eml_draft(path, row_values, rendered_html, row_number):
    message = EmailMessage()
    
    # Extract email headers from rendered HTML template
    headers = extract_email_headers_from_rendered_html(rendered_html)
    body_html = strip_email_headers_from_html(rendered_html)
    
    # Use extracted headers if available, otherwise fall back to row_values
    subject = headers.get('subject') or row_values.get("TITLE") or row_values.get("APPLICATION_CODE") or f"Generated draft row {row_number}"
    message["Subject"] = subject
    
    to_value = headers.get('to') or row_values.get("EMAIL_TO", "")
    cc_value = headers.get('cc') or row_values.get("EMAIL_CC", "")
    
    if to_value:
        message["To"] = to_value
    if cc_value:
        message["Cc"] = cc_value
    
    message.set_content("This is an HTML email draft generated by Doc Gen Studio.")
    message.add_alternative(body_html, subtype="html")
    path.write_text(message.as_string(), encoding="utf-8")


def create_msg_draft_via_outlook(path, row_values, rendered_html, row_number, report_stage=None):
    def stage(value):
        if report_stage:
            report_stage(value)

    try:
        stage("import pythoncom/win32com")
        import pythoncom  # type: ignore
        import win32com.client  # type: ignore
    except Exception as exc:
        raise RuntimeError("Outlook MSG output requires pywin32 and Microsoft Outlook on Windows.") from exc

    stage("CoInitialize")
    pythoncom.CoInitialize()
    outlook = None
    mail = None
    try:
        stage("Dispatch Outlook.Application")
        outlook = win32com.client.Dispatch("Outlook.Application")
        stage("CreateItem")
        mail = outlook.CreateItem(0)
        stage("set fields")
        
        # Extract email headers from rendered HTML template
        headers = extract_email_headers_from_rendered_html(rendered_html)
        body_html = strip_email_headers_from_html(rendered_html)
        
        # Use extracted headers if available, otherwise fall back to row_values
        mail.Subject = headers.get('subject') or row_values.get("TITLE") or row_values.get("APPLICATION_CODE") or f"Generated draft row {row_number}"
        
        # Use Recipients collection to properly handle email addresses
        to_value = headers.get('to') or row_values.get("EMAIL_TO", "")
        cc_value = headers.get('cc') or row_values.get("EMAIL_CC", "")
        
        if to_value:
            recipients = mail.Recipients
            recipients.Add(to_value)
            recipients(1).Type = 1  # olTo = 1
        
        if cc_value:
            recipients = mail.Recipients
            recipients.Add(cc_value)
            recipients(recipients.Count).Type = 2  # olCc = 2
        
        mail.HTMLBody = body_html
        
        stage("SaveAs")
        mail.SaveAs(str(path.resolve()), 9)
        stage("saved")
    except Exception as exc:
        raise RuntimeError(f"Could not create Outlook MSG draft: {exc}") from exc
    finally:
        mail = None
        outlook = None
        stage("CoUninitialize")
        pythoncom.CoUninitialize()


def self_worker_command():
    if getattr(sys, "frozen", False):
        return [sys.executable]
    return [sys.executable, str(Path(__file__).resolve())]


def run_outlook_msg_worker(payload_path):
    payload = json.loads(Path(payload_path).read_text(encoding="utf-8"))
    stage_path = Path(payload["stage_path"])

    def report_stage(value):
        stage_path.write_text(value, encoding="utf-8")

    create_msg_draft_via_outlook(
        Path(payload["path"]),
        {str(key): str(value) for key, value in payload["row_values"].items()},
        str(payload["rendered_html"]),
        int(payload["row_number"]),
        report_stage,
    )
    return 0


def write_msg_draft(path, row_values, rendered_html, row_number):
    with tempfile.TemporaryDirectory(prefix="docgen-msg-") as temp_dir:
        temp_path = Path(temp_dir)
        payload_path = temp_path / "payload.json"
        stage_path = temp_path / "stage.txt"
        payload = {
            "path": str(path.resolve()),
            "rendered_html": rendered_html,
            "row_number": row_number,
            "row_values": row_values,
            "stage_path": str(stage_path),
        }
        payload_path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")

        command = [*self_worker_command(), "--outlook-msg-worker", str(payload_path)]
        try:
            result = subprocess.run(
                command,
                capture_output=True,
                text=True,
                timeout=OUTLOOK_MSG_TIMEOUT_SECONDS,
                check=False,
            )
        except subprocess.TimeoutExpired as exc:
            last_stage = stage_path.read_text(encoding="utf-8") if stage_path.exists() else "startup"
            raise RuntimeError(
                f"Timed out after {OUTLOOK_MSG_TIMEOUT_SECONDS}s while creating Outlook MSG draft "
                f"for row {row_number} ({last_stage}). Close Outlook completely, reopen it, and try again."
            ) from exc

        if result.returncode != 0:
            details = (result.stderr or result.stdout).strip()
            raise RuntimeError(details or f"Outlook MSG worker exited with code {result.returncode}.")


def write_report(path, payload, placeholders, generated_rows, skipped_rows):
    rejection_rules = [
        f"{str(column).upper()} equals {', '.join(str(item) for item in values)}"
        for column, values in sorted(payload.get("skip_if_column_equals", {}).items())
    ]
    lines = [
        "Email Draft Generation Report",
        "",
        f"Workbook: {payload['workbook_path']}",
        f"Worksheet: {payload['worksheet_name']}",
        f"Output directory: {payload['output_dir']}",
        "",
        f"Email placeholders: {', '.join(sorted(placeholders)) or '(none)'}",
        f"Rejection rules: {'; '.join(rejection_rules) if rejection_rules else '(none)'}",
        f"Generated rows: {len(generated_rows)}",
        f"Skipped rows: {len(skipped_rows)}",
        "",
        "Skipped rows:",
    ]

    if skipped_rows:
        for row_number, missing in skipped_rows:
            lines.append(f"- Row {row_number}: missing {', '.join(missing)}")
    else:
        lines.append("- None")

    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main():
    if len(sys.argv) > 2 and sys.argv[1] == "--outlook-msg-worker":
        return run_outlook_msg_worker(sys.argv[2])

    payload = json.loads(sys.argv[1])
    emit_progress("prepare", "Loading workbook and email template.", percent=5)

    workbook = load_workbook(payload["workbook_path"], data_only=True)
    worksheet_name = (payload.get("worksheet_name") or "").strip()
    if not worksheet_name:
        worksheet_name = workbook.sheetnames[0]
    if worksheet_name not in workbook.sheetnames:
        raise ValueError(
            f"Worksheet {worksheet_name!r} not found. Available sheets: {', '.join(workbook.sheetnames)}"
        )

    worksheet = workbook[worksheet_name]
    data_start_row = int(payload["data_start_row"])
    template_text = payload["email_template_text"]
    email_output_mode = payload.get("email_output_mode", "combined_docx")
    mapping = payload["mapping"]
    skip_if_column_equals = {
        str(column).upper(): [str(item) for item in values]
        for column, values in payload.get("skip_if_column_equals", {}).items()
    }
    output_dir = Path(payload["output_dir"])
    output_dir.mkdir(parents=True, exist_ok=True)

    placeholders = set(PLACEHOLDER_RE.findall(template_text))
    generated_rows = []
    skipped_rows = []
    drafts = []
    separate_drafts = []
    total_rows = max(0, worksheet.max_row - data_start_row + 1)
    emit_progress(
        "prepare",
        f"{total_rows} workbook rows found.",
        total=total_rows,
        percent=10,
        rowsFound=total_rows,
    )

    for row_index, row_number in enumerate(range(data_start_row, worksheet.max_row + 1), start=1):
        emit_progress(
            "email",
            f"Preparing email draft for row {row_number}.",
            current=row_index,
            total=total_rows,
            percent=stage_percent(15, 85, row_index, total_rows),
            emailDraftCount=len(generated_rows),
            generatedCount=len(generated_rows),
            rowsFound=total_rows,
            skippedCount=len(skipped_rows),
        )
        if row_matches_skip_rules(worksheet, row_number, skip_if_column_equals):
            skipped_rows.append((row_number, ["rejection value"]))
            emit_progress(
                "email",
                f"Skipped workbook row {row_number}.",
                current=row_index,
                total=total_rows,
                percent=stage_percent(15, 85, row_index, total_rows),
                emailDraftCount=len(generated_rows),
                generatedCount=len(generated_rows),
                rowsFound=total_rows,
                skippedCount=len(skipped_rows),
            )
            continue

        row_values = build_row_values(worksheet, row_number, mapping)
        missing = sorted(
            placeholder for placeholder in placeholders if not row_values.get(placeholder, "")
        )

        if missing:
            skipped_rows.append((row_number, missing))
            emit_progress(
                "email",
                f"Skipped workbook row {row_number}.",
                current=row_index,
                total=total_rows,
                percent=stage_percent(15, 85, row_index, total_rows),
                emailDraftCount=len(generated_rows),
                generatedCount=len(generated_rows),
                rowsFound=total_rows,
                skippedCount=len(skipped_rows),
            )
            continue

        rendered = render_template(template_text, row_values)
        application_code = row_values.get("APPLICATION_CODE", "")
        heading = (
            f"===== {application_code} - ROW {row_number} ====="
            if application_code
            else f"===== ROW {row_number} ====="
        )
        drafts.append((heading, rendered))
        separate_drafts.append((row_number, heading, rendered, row_values))
        generated_rows.append(row_number)
        emit_progress(
            "email",
            f"Prepared email draft for row {row_number}.",
            current=row_index,
            total=total_rows,
            percent=stage_percent(15, 85, row_index, total_rows),
            emailDraftCount=len(generated_rows),
            generatedCount=len(generated_rows),
            rowsFound=total_rows,
            skippedCount=len(skipped_rows),
        )

    combined_email_path = output_dir / "email_drafts.html"
    email_output_dir = output_dir / "emails"
    report_path = output_dir / "generation_report.txt"

    emit_progress(
        "report",
        "Writing email draft files and report.",
        percent=92,
        emailDraftCount=len(generated_rows),
        expectedEmailDraftCount=len(generated_rows),
        generatedCount=len(generated_rows),
        rowsFound=total_rows,
        skippedCount=len(skipped_rows),
    )
    if email_output_mode == "combined_docx":
        write_email_drafts_html(combined_email_path, drafts)
    elif email_output_mode in {"separate_docx", "separate_eml", "separate_msg"}:
        email_output_dir.mkdir(parents=True, exist_ok=True)
        email_filename_counter = Counter()
        for row_number, heading, rendered, row_values in separate_drafts:
            application_code = row_values.get("APPLICATION_CODE", "")
            draft_base_name = (
                f"email draft - {application_code}"
                if application_code
                else f"email draft row {row_number}"
            )

            if email_output_mode == "separate_docx":
                html_filename = ensure_unique_filename(draft_base_name, ".html", email_filename_counter, row_number)
                (email_output_dir / html_filename).write_text(
                    wrap_single_email_html(rendered, heading),
                    encoding="utf-8",
                )
            else:
                if email_output_mode == "separate_eml":
                    eml_filename = ensure_unique_filename(draft_base_name, ".eml", email_filename_counter, row_number)
                    write_eml_draft(email_output_dir / eml_filename, row_values, rendered, row_number)
                else:
                    msg_filename = ensure_unique_filename(draft_base_name, ".msg", email_filename_counter, row_number)
                    write_msg_draft(email_output_dir / msg_filename, row_values, rendered, row_number)

    write_report(report_path, payload, placeholders, generated_rows, skipped_rows)

    print(f"Generated {len(generated_rows)} emails.")
    print(f"Skipped {len(skipped_rows)} rows.")
    if email_output_mode == "combined_docx":
        print(f"Combined email drafts file: {combined_email_path}")
    else:
        print(f"Email drafts directory: {email_output_dir}")
    print(f"Report: {report_path}")
    emit_progress(
        "complete",
        "Email draft generation complete.",
        percent=100,
        emailDraftCount=len(generated_rows),
        expectedEmailDraftCount=len(generated_rows),
        generatedCount=len(generated_rows),
        rowsFound=total_rows,
        skippedCount=len(skipped_rows),
    )


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"Error: {exc}", file=sys.stderr)
        raise
