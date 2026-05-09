import json
import html
import re
import string
import sys
import unicodedata
from pathlib import Path
from zipfile import ZipFile

from openpyxl import load_workbook


PLACEHOLDER_RE = re.compile(r"\{\{([A-Z0-9_]+)\}\}")
XML_PART_PREFIXES = (
    "word/document.xml",
    "word/header",
    "word/footer",
    "word/footnotes.xml",
    "word/endnotes.xml",
)

def normalize_cell_value(value):
    if value is None:
        return ""
    return str(value).strip()


HEADER_VARIABLE_ALIASES = [
    ("APPLICATION_CODE", ["#", "application code", "application", "code", "κωδικός αίτησης", "κωδικος αιτησης", "κωδικός", "κωδικος", "kodikos aitisis", "kodikos"]),
    ("EMAIL_TO", ["email to", "to", "email", "προς", "email προς", "paraliptis", "pros"]),
    ("EMAIL_CC", ["cc", "κοινοποίηση", "κοινοποιηση", "cc email", "koinopoiisi"]),
    ("AUTHOR", ["author", "συγγραφέας", "συγγραφεας", "syggrafeas", "singrafeas"]),
    ("PUBLISHER", ["publisher", "εκδότης", "εκδοτης", "ekdotis"]),
    ("AMOUNT", ["amount", "ποσό", "ποσο", "poso"]),
    ("FIRST_INSTALLMENT", ["first installment", "installment", "πρώτη δόση", "πρωτη δοση", "dosi", "proti dosi"]),
    ("LANGUAGE", ["language", "γλώσσα", "γλωσσα", "glossa", "glwssa"]),
    ("TITLE", ["title", "τίτλος", "τιτλος", "titlos"]),
]

MAX_SUGGESTED_VARIABLE_LENGTH = 24

GREEK_TO_LATIN = str.maketrans({
    "α": "a", "β": "v", "γ": "g", "δ": "d", "ε": "e", "ζ": "z", "η": "i", "θ": "th",
    "ι": "i", "κ": "k", "λ": "l", "μ": "m", "ν": "n", "ξ": "x", "ο": "o", "π": "p",
    "ρ": "r", "σ": "s", "ς": "s", "τ": "t", "υ": "y", "φ": "f", "χ": "ch", "ψ": "ps", "ω": "o",
    "Α": "A", "Β": "V", "Γ": "G", "Δ": "D", "Ε": "E", "Ζ": "Z", "Η": "I", "Θ": "TH",
    "Ι": "I", "Κ": "K", "Λ": "L", "Μ": "M", "Ν": "N", "Ξ": "X", "Ο": "O", "Π": "P",
    "Ρ": "R", "Σ": "S", "Τ": "T", "Υ": "Y", "Φ": "F", "Χ": "CH", "Ψ": "PS", "Ω": "O",
})


def strip_accents(value):
    decomposed = unicodedata.normalize("NFD", value)
    return "".join(char for char in decomposed if unicodedata.category(char) != "Mn")


def normalize_header_text(value):
    without_accents = strip_accents(value)
    return re.sub(r"\s+", " ", without_accents).strip().lower()


def greek_to_latin(value):
    return strip_accents(value).translate(GREEK_TO_LATIN)


def alias_matches(value, alias):
    normalized_alias = normalize_header_text(alias)
    if not normalized_alias:
        return False
    if value == normalized_alias:
        return True
    if len(normalized_alias) < 4:
        return False
    return re.search(rf"(?<![a-z0-9]){re.escape(normalized_alias)}(?![a-z0-9])", value) is not None


def compact_variable_name(value, max_length=MAX_SUGGESTED_VARIABLE_LENGTH):
    parts = [part for part in re.split(r"[^A-Za-z0-9]+", value) if part]
    selected = []
    current_length = 0

    for part in parts:
        next_length = current_length + len(part) + (1 if selected else 0)
        if selected and next_length > max_length:
            break
        if not selected and len(part) > max_length:
            selected.append(part[:max_length])
            break

        selected.append(part)
        current_length = next_length

    return "_".join(selected).upper() or None


def normalize_header_to_variable(header):
    header_value = normalize_cell_value(header)
    if not header_value:
        return None

    normalized = normalize_header_text(header_value)
    normalized_latin = normalize_header_text(greek_to_latin(header_value))

    for variable, aliases in HEADER_VARIABLE_ALIASES:
        normalized_aliases = {normalize_header_text(alias) for alias in aliases}
        if normalized in normalized_aliases or normalized_latin in normalized_aliases:
            return variable

    for variable, aliases in HEADER_VARIABLE_ALIASES:
        for alias in aliases:
            if alias_matches(normalized, alias) or alias_matches(normalized_latin, alias):
                return variable

    fallback_source = greek_to_latin(header_value)
    return compact_variable_name(fallback_source)



def get_column_letter(index):
    letters = []
    while index > 0:
        index, remainder = divmod(index - 1, 26)
        letters.append(string.ascii_uppercase[remainder])
    return "".join(reversed(letters))


def get_column_index(column_letter):
    index = 0
    for letter in column_letter.upper():
        if letter not in string.ascii_uppercase:
            return None
        index = index * 26 + string.ascii_uppercase.index(letter) + 1
    return index or None


def extract_contract_tokens(path_value):
    if not path_value:
        return {"contexts": {}, "tokens": []}

    path = Path(path_value)
    if not path.exists() or path.suffix.lower() != ".docx":
        return {"contexts": {}, "tokens": []}

    tokens = set()
    token_contexts = {}

    def extract_paragraph_texts(content):
        paragraphs = re.findall(r"<w:p(?:\s[^>]*)?>(.*?)</w:p>", content, flags=re.DOTALL)
        paragraph_texts = []

        for paragraph in paragraphs:
            text_nodes = re.findall(r"<w:t(?:\s[^>]*)?>(.*?)</w:t>", paragraph, flags=re.DOTALL)
            if not text_nodes:
                continue
            text = "".join(html.unescape(node) for node in text_nodes)
            text = re.sub(r"\s+", " ", text).strip()
            if text:
                paragraph_texts.append(text)

        return paragraph_texts

    with ZipFile(path, "r") as archive:
        for name in archive.namelist():
            if not any(name == prefix or name.startswith(prefix) for prefix in XML_PART_PREFIXES):
                continue
            content = archive.read(name).decode("utf-8", errors="ignore")

            for paragraph_text in extract_paragraph_texts(content):
                paragraph_tokens = PLACEHOLDER_RE.findall(paragraph_text)
                if not paragraph_tokens:
                    continue

                for token in paragraph_tokens:
                    tokens.add(token)
                    token_contexts.setdefault(token, paragraph_text)

            tokens.update(PLACEHOLDER_RE.findall(content))

    sorted_tokens = sorted(tokens)
    filtered_contexts = {
        token: token_contexts[token]
        for token in sorted_tokens
        if token in token_contexts
    }
    return {
        "contexts": filtered_contexts,
        "tokens": sorted_tokens,
    }


def parse_positive_int(value, fallback):
    try:
        parsed = int(value)
        return parsed if parsed > 0 else fallback
    except (TypeError, ValueError):
        return fallback


def count_non_empty_cells(row):
    return sum(1 for cell in row if normalize_cell_value(cell.value))


def analyze_header_rows(worksheet, selected_header_row, scan_limit=10):
    max_scan_row = min(worksheet.max_row or selected_header_row, scan_limit)
    selected_header_count = 0
    best_row = selected_header_row
    best_count = 0

    for row_number in range(1, max_scan_row + 1):
        row = worksheet[row_number]
        non_empty_count = count_non_empty_cells(row)
        if row_number == selected_header_row:
            selected_header_count = non_empty_count
        if non_empty_count > best_count:
            best_count = non_empty_count
            best_row = row_number

    suggested_header_row = None
    suggested_header_count = 0
    if best_row != selected_header_row and best_count >= max(2, selected_header_count + 2):
        suggested_header_row = best_row
        suggested_header_count = best_count

    return {
        "scannedRows": max_scan_row,
        "selectedHeaderCount": selected_header_count,
        "suggestedHeaderCount": suggested_header_count,
        "suggestedHeaderRow": suggested_header_row,
    }


def preview_row_cells(worksheet, row_number, required_letters=None, max_columns=8):
    required_letters = required_letters or []
    required_indexes = {
        index
        for index in (get_column_index(letter) for letter in required_letters)
        if index is not None
    }
    visible_indexes = set(range(1, min(worksheet.max_column or 1, max_columns) + 1))
    visible_indexes.update(required_indexes)
    cells = []
    for index in sorted(visible_indexes):
        cell = worksheet.cell(row=row_number, column=index)
        cells.append({
            "columnLetter": get_column_letter(cell.column),
            "value": normalize_cell_value(cell.value),
        })
    return cells


def find_rejected_preview_row(worksheet, data_start_row, rejection_column, rejection_value):
    if not rejection_column or not rejection_value:
        return None

    rejection_index = get_column_index(rejection_column)
    if rejection_index is None:
        return None

    for row_number in range(data_start_row, (worksheet.max_row or data_start_row) + 1):
        cell_value = normalize_cell_value(worksheet.cell(row=row_number, column=rejection_index).value)
        if cell_value == rejection_value:
            return row_number

    return None


def build_preview_rows(
    worksheet,
    header_row,
    data_start_row,
    header_analysis,
    required_letters=None,
    rejection_column=None,
    rejection_value=None,
):
    preview_row_numbers = [header_row]
    suggested_header_row = header_analysis.get("suggestedHeaderRow")
    if suggested_header_row and suggested_header_row not in preview_row_numbers:
        preview_row_numbers.append(suggested_header_row)
    for row_number in range(data_start_row, min(worksheet.max_row or data_start_row, data_start_row + 2) + 1):
        if row_number not in preview_row_numbers:
            preview_row_numbers.append(row_number)
    rejected_preview_row = find_rejected_preview_row(
        worksheet,
        data_start_row,
        rejection_column,
        rejection_value,
    )
    if rejected_preview_row and rejected_preview_row not in preview_row_numbers:
        preview_row_numbers.append(rejected_preview_row)

    preview_rows = []
    sorted_row_numbers = sorted(preview_row_numbers)
    previous_row_number = None
    for row_number in sorted_row_numbers:
        if row_number == header_row:
            role = "selected-header"
        elif row_number == suggested_header_row:
            role = "suggested-header"
        elif row_number == rejected_preview_row:
            role = "rejected-data"
        else:
            role = "data"

        row = worksheet[row_number]
        preview_rows.append({
            "cells": preview_row_cells(worksheet, row_number, required_letters),
            "hasLeadingGap": previous_row_number is not None and row_number > previous_row_number + 1,
            "populatedCells": count_non_empty_cells(row),
            "role": role,
            "rowNumber": row_number,
        })
        previous_row_number = row_number

    return preview_rows


def inspect_workbook(request_payload):
    workbook_path = (request_payload.get("workbook_path") or "").strip()
    if not workbook_path:
        contract_info = extract_contract_tokens(request_payload.get("contractTemplatePath"))
        return {
            "columnValues": {},
            "columns": [],
            "contractTokenContexts": contract_info["contexts"],
            "contractTokens": contract_info["tokens"],
            "dataStartRow": 2,
            "headerAnalysis": None,
            "headerRow": 1,
            "maxColumn": 0,
            "previewRows": [],
            "sampleRows": [],
            "totalRows": 0,
            "skippedRows": 0,
            "worksheetName": "",
            "worksheetNames": [],
        }
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    worksheet_names = list(workbook.sheetnames)
    worksheet_name = (request_payload.get("worksheet_name") or "").strip()
    if worksheet_name and worksheet_name in worksheet_names:
        worksheet = workbook[worksheet_name]
    else:
        worksheet = workbook[worksheet_names[0]]
    header_row = parse_positive_int(request_payload.get("header_row"), 1)
    data_start_row = parse_positive_int(request_payload.get("data_start_row"), max(2, header_row + 1))
    header_analysis = analyze_header_rows(worksheet, header_row)

    columns = []
    rejection_column = (request_payload.get("rejectionColumn") or "").strip().upper()
    rejection_value = (request_payload.get("rejectionValue") or "").strip()
    preview_rows = build_preview_rows(
        worksheet,
        header_row,
        data_start_row,
        header_analysis,
        [rejection_column] if rejection_column else [],
        rejection_column,
        rejection_value,
    )

    columns = []
    for cell in worksheet[header_row]:
        if cell.value is None:
            continue
        header = normalize_cell_value(cell.value)
        if not header:
            continue
        sample_value = normalize_cell_value(
            worksheet[f"{get_column_letter(cell.column)}{data_start_row}"].value
        )
        columns.append(
            {
                "columnLetter": get_column_letter(cell.column),
                "header": header,
                "sampleValue": sample_value,
                "suggestedVariable": normalize_header_to_variable(header),
            }
        )

    # Build a letter->index map so we can pick the right cell from each row cheaply
    col_letter_to_idx = {column["columnLetter"]: i for i, column in enumerate(columns)}
    column_letters = [column["columnLetter"] for column in columns]
    column_values = {letter: [] for letter in column_letters}
    seen_column_values = {letter: set() for letter in column_letters}
    MAX_UNIQUE_VALUES = 200
    rejection_col_idx = col_letter_to_idx.get(rejection_column)
    skipped_rows = 0
    for row in worksheet.iter_rows(min_row=data_start_row, values_only=True):
        if rejection_col_idx is not None and rejection_value:
            cell_val = normalize_cell_value(row[rejection_col_idx] if rejection_col_idx < len(row) else None)
            if cell_val == rejection_value:
                skipped_rows += 1
                continue
        for letter in column_letters:
            idx = col_letter_to_idx[letter]
            if idx >= len(row):
                continue
            if len(column_values[letter]) >= MAX_UNIQUE_VALUES:
                continue
            value = normalize_cell_value(row[idx])
            if value in seen_column_values[letter]:
                continue
            seen_column_values[letter].add(value)
            column_values[letter].append(value)

    sample_rows = []
    for row_idx, row in enumerate(worksheet.iter_rows(min_row=data_start_row, max_row=data_start_row + 2, values_only=True)):
        row_values = {}
        for column in columns[:6]:
            idx = col_letter_to_idx[column["columnLetter"]]
            row_values[column["header"]] = normalize_cell_value(row[idx] if idx < len(row) else None)
        sample_rows.append(
            {
                "rowNumber": data_start_row + row_idx,
                "values": row_values,
            }
        )

    workbook.close()
    contract_info = extract_contract_tokens(request_payload.get("contractTemplatePath"))

    return {
        "columnValues": column_values,
        "columns": columns,
        "contractTokenContexts": contract_info["contexts"],
        "contractTokens": contract_info["tokens"],
        "dataStartRow": data_start_row,
        "headerAnalysis": header_analysis,
        "headerRow": header_row,
        "maxColumn": worksheet.max_column or 0,
        "previewRows": preview_rows,
        "sampleRows": sample_rows,
        "totalRows": max(0, worksheet.max_row - data_start_row + 1),
        "skippedRows": skipped_rows,
        "worksheetName": worksheet.title,
        "worksheetNames": worksheet_names,
    }


if __name__ == "__main__":
    raw_payload = json.loads(sys.argv[1])
    payload = {
        "contractTemplatePath": raw_payload.get("contractTemplatePath"),
        "data_start_row": raw_payload.get("dataStartRow", 2),
        "header_row": raw_payload.get("headerRow", 1),
        "workbook_path": raw_payload.get("workbookPath", ""),
        "worksheet_name": raw_payload.get("worksheetName"),
        "rejectionColumn": raw_payload.get("rejectionColumn"),
        "rejectionValue": raw_payload.get("rejectionValue"),
    }
    print(json.dumps(inspect_workbook(payload)))
